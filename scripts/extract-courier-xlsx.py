#!/usr/bin/env python3
"""Extract courier quoting reference data from the three West Point workbooks.

Usage:
    python3 scripts/extract-courier-xlsx.py <dir-with-workbooks> [out-dir]

Reads:
    *BPP_Element_Estimator.xlsm      -> reference tables (towns, zones, SLA, line-haul)
    *DSV_Rates.xlsx                  -> DSV rate card (long form) + metadata
    *Courier_Comparison_v2.xlsm      -> MJV area rates + city classifications

Writes CSVs to data/courier/ plus golden test fixtures extracted from the
estimator's INPUT sheet (30 worked examples with expected Element + SLA).
"""
import csv
import glob
import json
import os
import re
import sys

import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "."
OUT = sys.argv[2] if len(sys.argv) > 2 else "data/courier"
os.makedirs(OUT, exist_ok=True)


def find(pattern):
    hits = glob.glob(os.path.join(SRC, f"*{pattern}*"))
    if not hits:
        sys.exit(f"missing workbook matching *{pattern}* in {SRC}")
    return hits[0]


def norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    return re.sub(r"\s+", " ", s)


def write_csv(name, header, rows):
    path = os.path.join(OUT, name)
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"{name}: {len(rows)} rows")


def sheet_rows(ws, cols, start_row, key_col=0):
    """Extract given column indexes (1-based) from start_row; skip rows whose key is blank."""
    out = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        vals = [norm(row[c - 1]) if c - 1 < len(row) else "" for c in cols]
        if vals[key_col]:
            out.append(vals)
    return out


# ---------------------------------------------------------------- estimator
bpp = openpyxl.load_workbook(find("BPP_Element_Estimator"), data_only=True, read_only=True)

ws = bpp["MB_TownList"]
write_csv(
    "mb_townlist.csv",
    ["town_code", "tariff_hub", "ops_branch", "town_name", "province", "town_type",
     "town_type_abbr", "remote_km", "servicing_town", "service_code", "location",
     "sla_facility", "zone"],
    sheet_rows(ws, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 13, 14], 2),
)

ws = bpp["MB_Suburblist"]
hdr = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
write_csv(
    "mb_suburblist.csv",
    ["suburb", "col2", "col3", "col4", "town_code"],
    sheet_rows(ws, [1, 2, 3, 4, 5], 2),
)

ws = bpp["BPP_FInancialTownlist"]
write_csv(
    "bpp_financial_townlist.csv",
    ["from_date", "fin_network", "division", "branch_col", "service_level", "tarsrt",
     "country", "city", "city2", "from_postcode", "thru_postcode", "till_date",
     "zone_code", "remark", "lookup", "branch", "zone"],
    sheet_rows(ws, list(range(1, 18)), 2, key_col=7),
)

ws = bpp["Townslist"]
write_csv(
    "townslist.csv",
    ["town_name", "billing_town", "zone", "ops_branch"],
    sheet_rows(ws, [1, 2, 3, 4], 2),
)

ws = bpp["LH&Terminal"]
rows_lh, rows_direct = [], []
for row in ws.iter_rows(min_row=2, values_only=True):
    a = [norm(v) for v in (list(row) + [""] * 8)[:8]]
    if a[0]:
        rows_lh.append(a[:4])       # OrigBranch, DestBranch, LHProductDesc, Total
    if a[6]:
        rows_direct.append(a[6:8])  # direct-override lookup key, element
write_csv("lh_terminal.csv", ["orig_branch", "dest_branch", "lh_product_desc", "total_legs"], rows_lh)
write_csv("lh_direct_overrides.csv", ["lookup_key", "element"], rows_direct)

ws = bpp["SLA lookups"]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    a = [norm(v) for v in (list(row) + [""] * 25)[:25]]
    if any(a):
        rows.append(a)
write_csv(
    "sla_lookups.csv",
    ["orig_type", "dest_type", "type_lookup", "sla_type", "sla_type_code",
     "route", "tariff", "c8", "service", "service_code", "facility_type",
     "route_band", "route_type", "lookup", "sla_days", "sla_text", "c17",
     "blns_service", "blns_band", "blns_code", "blns_lookup", "blns_sla",
     "onx_lookup", "onx_days", "onx_sla"],
    rows,
)

ws = bpp["Sheet4"]
write_csv("branch_map.csv", ["ops_branch", "tariff_branch"], sheet_rows(ws, [1, 2], 2))

control = norm(bpp["Control"]["B2"].value)
print(f"rate-card expiry guard (Control!B2): {control}")

# golden fixtures from INPUT
ws = bpp["INPUT"]
fixtures = []
for row in ws.iter_rows(min_row=4, values_only=True):
    r = [norm(v) for v in (list(row) + [""] * 29)[:29]]
    origin, opc, dest, dpc, service, element = r[1], r[2], r[3], r[4], r[5], r[6]
    if not origin:
        continue
    fixtures.append({
        "origin": origin, "originPostcode": opc,
        "destination": dest, "destPostcode": dpc,
        "service": service,
        "expected": {
            "element": element, "sla": r[11],
            "puZone": r[12], "delZone": r[13],
            "opsOrigin": r[14], "opsDest": r[15], "balance": r[16],
            "origTownCode": r[17], "destTownCode": r[18],
            "blns": r[22], "slaType": r[23],
            "oTariffHub": r[24], "dTariffHub": r[25],
        },
    })
with open(os.path.join(OUT, "golden_fixtures.json"), "w") as f:
    json.dump({"controlExpiry": control, "cases": fixtures}, f, indent=2)
print(f"golden_fixtures.json: {len(fixtures)} cases")

# ---------------------------------------------------------------- DSV rate cards
# The workbook stacks two grids in one sheet: an Economy card and an Express
# card, each headed "Account Name / Representative / Date / <Service> Rate".
dsv = openpyxl.load_workbook(find("DSV_Rates"), data_only=True, read_only=True)
ws = dsv["Table 1"]
grid = [[norm(v) for v in row] for row in ws.iter_rows(values_only=True)]

service_starts = [i for i, g in enumerate(grid) if g and re.match(r"(Economy|Express) Rate", g[0] or "")]
all_meta = {}
for start in service_starts:
    service = grid[start][0].split()[0]  # Economy | Express
    meta = {"account": grid[start - 3][0], "representative": grid[start - 2][0],
            "effective_date": grid[start - 1][0]}
    band_row = next(g for g in grid[start:start + 6] if any(v and ("to <" in v or "Over" in v) for v in g))
    band_cols = [(c, v) for c, v in enumerate(band_row) if v and ("to <" in v or "Over" in v)]
    min_row = next(g for g in grid[start:start + 4] if g[0] == "Minimum")
    minimum = next((v for v in min_row if re.fullmatch(r"\d+(\.\d+)?", v)), "")

    rate_rows = []
    started = False
    for g in grid[start:]:
        if g[0] == "Elements":
            started = True
            continue
        if not started:
            continue
        if g[0] and re.fullmatch(r"\d+", g[0]) and int(g[0]) <= 30:
            nums = [v for v in g[1:] if re.fullmatch(r"\d+(\.\d+)?", v)]
            for (col, band), price in zip(band_cols, nums):
                rate_rows.append([int(g[0]), band, price])
        elif rate_rows:
            break  # end of this grid
    write_csv(f"dsv_rate_card_{service.lower()}.csv", ["element", "weight_band", "price"], rate_rows)
    all_meta[service] = {**meta, "minimum": minimum, "bands": [b for _, b in band_cols]}
    print(f"DSV {service} card: {meta}")

with open(os.path.join(OUT, "dsv_rate_card_meta.json"), "w") as f:
    json.dump(all_meta, f, indent=2)

# ---------------------------------------------------------------- comparison workbook
cmp_wb = openpyxl.load_workbook(find("Courier_Comparison"), data_only=True, read_only=True)
ws = cmp_wb["Lists"]
grid = [[norm(v) for v in row] for row in ws.iter_rows(values_only=True)]

area_rows = [g[:4] for g in grid[1:5] if g[0] in ("Main", "Local", "Outlying")]
write_csv("mjv_area_rates.csv", ["area", "rate_per_kg", "ta_per_kg", "fuel_levy"], area_rows)

city_rows = [[g[19], g[20]] for g in grid[1:] if len(g) > 20 and g[19] and g[20]]
write_csv("city_classifications.csv", ["city", "classification"], city_rows)

band_rows = [g[8:14] for g in grid[1:22] if g[8]]
write_csv("dsv_weight_bands.csv",
          ["band_floor_kg", "range_label", "rate_col", "per_unit", "min_rate", "fuel_levy"],
          band_rows)

fuel_rows = [g[:5] for g in grid[30:35] if g[0]]
write_csv("fuel_levy_history.csv",
          ["month", "mjv_levy", "mjv_change", "dsv_levy", "dsv_change"], fuel_rows)

print("\nDone. Output in", OUT)
